import calendar
import datetime
import sqlite3
import sys
from os.path import exists

import numpy as np
import openpyxl
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTime, QDate
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.page import PrintPageSetup

from DialogsDesigns.addFamilyMemberDialog import addFamilyMemberDialog
from DialogsDesigns.changeColorFamilyMemberDialog import changeColorFamilyMemberDialog
from DialogsDesigns.deleteEventDialog import deleteEventDialog
from DialogsDesigns.design import Ui_MainWindow as mainWindowDesign
from DialogsDesigns.removeFamilyMemberDialog import removeFamilyMemberDialog


class Window(QMainWindow, mainWindowDesign):
    def __init__(self):
        super().__init__()
        if exists("events.db"):
            self.con = sqlite3.connect("events.db")
        else:
            pass  # Сделать создание базы данных при её отсутствии TODO
        self.cur = self.con.cursor()
        self.setupUI()

    def setupUI(self):
        self.setupUi(self)
        self.setMinimumSize(1024, 768)
        self.checkBoxes = []
        self.typesOfRegular = {i[0]: i[1] for i in
                               self.cur.execute('''select * from typesOfRegular''')}
        self.stackedWidget.setCurrentIndex(0)
        self.updateFamilyMembersCheckBoxes()

        self.addEventButton.clicked.connect(self.onAddEventButtonClicked)
        self.deleteEventButton.clicked.connect(self.onDeleteEventButtonClicked)
        self.addFamilyMemberButton.clicked.connect(self.addFamilyMemberClicked)
        self.changeColorFamilyMember.clicked.connect(
            self.onChangeColorFamilyMemberClicked)
        self.removeFamilyMember.clicked.connect(self.onRemoveFamilyMemberClicked)
        self.toExcelButton.clicked.connect(self.toExcelButtonClicked)

        self.dateComboBox.addItems(self.typesOfRegular.values())
        self.dateComboBox.setCurrentIndex(len(self.typesOfRegular) - 1)
        self.addButton.clicked.connect(self.addEvent)
        self.cancelButton.clicked.connect(lambda: self.stackedWidget.setCurrentIndex(0))

    def updateDB(self):
        self.colors = list(self.cur.execute('''select * from colors'''))
        self.familyMembers = tuple(map(lambda x: x[1:], self.cur.execute(
            '''select * from familyMembers''')))
        self.events = tuple(self.cur.execute('''select * from events'''))
        self.calendarWidget.updateDB()

    # page 1
    def addFamilyMemberClicked(self):
        dlg = addFamilyMemberDialog()
        dlg.exec()
        if dlg.colorID is not None:
            self.updateFamilyMembersCheckBoxes()

    def onChangeColorFamilyMemberClicked(self):
        dlg = changeColorFamilyMemberDialog()
        dlg.exec()
        if dlg.resultLabel.text() != 'Выберите новый цвет':
            self.updateFamilyMembersCheckBoxes()

    def onRemoveFamilyMemberClicked(self):
        try:
            dlg = removeFamilyMemberDialog()
            dlg.exec()
            self.updateFamilyMembersCheckBoxes()
        except BaseException as e:
            print(e)

    def onAddEventButtonClicked(self):
        try:
            self.stackedWidget.setCurrentIndex(1)
            time_now = datetime.datetime.now().time()
            date_now = datetime.datetime.now().date()
            self.dateTimeEdit.setTime((QTime(time_now.hour, time_now.minute)))
            self.dateTimeEdit.setDate((QDate(date_now.year, date_now.month, date_now.day)))
            self.familyMembersComboBox.clear()
            self.familyMembersComboBox.addItems(tuple(map(lambda x: x[0], self.familyMembers)))
        except BaseException as e:
            print(e)

    def updateFamilyMembersCheckBoxes(self):
        self.updateDB()
        for checkbox in self.checkBoxes:
            self.verticalLayout.removeWidget(checkbox)
        self.checkBoxes.clear()
        for (memberName, colorID) in sorted(self.familyMembers, key=lambda x: x[0]):
            self.checkBoxes.append(QtWidgets.QCheckBox(self.centralwidget))
            self.checkBoxes[-1].setText(memberName)
            self.checkBoxes[-1].setChecked(True)
            r, g, b = list(filter(lambda x: x[0] == colorID, self.colors))[0][1:]
            new_css = f'''color: rgb(255, 255, 255);
background-color: rgb({r}, {g}, {b});''' if r == g == b == 0 else \
                f'''color: rgb(0, 0, 0);
background-color: rgb({r}, {g}, {b});'''
            self.checkBoxes[-1].setStyleSheet(new_css + '''border-style: outset;
border-width: 2px;
border-radius: 10px;
border-color: rgb(0, 0, 0);
font: bold 14px;
min-width: 10em;
padding: 6px;''')
            self.verticalLayout.insertWidget(0, self.checkBoxes[-1])

    def toExcelButtonClicked(self):
        month_array = get_month_array(self.calendarWidget.yearShown(),
                                      self.calendarWidget.monthShown())
        month = self.calendarWidget.monthShown()

        month_title = {1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
                       7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь",
                       12: "Декабрь"}[month]
        events = tuple(
            sqlite3.connect("events.db").cursor().execute('''select title, date from events'''))
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            font = Font(size=8)
            for i, element in enumerate("ABCDEFG"):
                ws.column_dimensions[element].width = 20
            for i in range(len(month_array) + 1):
                ws.row_dimensions[i + 1].height = 40
            for i, element in enumerate(
                    ("Понедельник", "Вторник", "Среда", "Четверг", "Пятница",
                     "Суббота", "Воскресенье")):
                ws[f'{"ABCDEFG"[i]}1'] = element
                ws[f'{"ABCDEFG"[i]}1'].border = thin_border
                ws[f'{"ABCDEFG"[i]}1'].font = font
                ws[f'{"ABCDEFG"[i]}1'].alignment = Alignment(horizontal='center',
                                                             vertical='center')

            for i, line in enumerate(month_array):
                for j, element in enumerate(line):
                    final_text = str(element) if element else ''
                    for title, date in events:
                        if date.split('.')[0] == str(element) and date.split('.')[1] == str(
                                month):
                            final_text += f"\n{title}"
                    ws[f'{"ABCDEFGHIJKLMNOPQRSTUVWXYZ"[j]}{i + 2}'] = final_text
                    cell = ws[f'{"ABCDEFGHIJKLMNOPQRSTUVWXYZ"[j]}{i + 2}']
                    cell.alignment = Alignment(wrapText=True, horizontal='left', vertical='top')
                    cell.border = thin_border
                    cell.font = font
            ws_properties = ws.sheet_properties
            ws.page_setup = PrintPageSetup(orientation='landscape')
            wb.save(f"{month_title}.xlsx")
        except BaseException as e:
            print(e)

    def onDeleteEventButtonClicked(self):
        try:
            dlg = deleteEventDialog()
            dlg.exec()
            self.updateDB()
            self.calendarWidget.updateDB()
            self.calendarWidget.repaint()
        except BaseException as e:
            print(e)

    # page 2
    def addEvent(self):
        if self.titleEdit.text() == '':
            dlg = QMessageBox(self)
            dlg.setWindowTitle('Ошибка')
            dlg.setText('Не указан заголовок события!')
            dlg.exec()
        else:
            try:
                f_m = tuple(self.cur.execute(f'''select id from familyMembers 
                where name = "{self.familyMembersComboBox.currentText()}"'''))[0][0]
                t_o_r = tuple(self.cur.execute(f'''select id from typesOfRegular 
                where title = "{self.dateComboBox.currentText()}"'''))[0][0]
                date = self.dateTimeEdit.date()
                time = self.dateTimeEdit.time()
                self.cur.execute(
                    f'''insert into events values({f_m}, {t_o_r}, "{self.titleEdit.text()}", 
                    "{self.textEdit.text()}", "{date.day()}.{date.month()}.{date.year()}",
                    "{time.hour()}:{time.minute()}")''')
                self.con.commit()
                self.stackedWidget.setCurrentIndex(0)
                self.calendarWidget.repaint()
            except BaseException as e:
                print(e)


def get_month_array(year, month):
    try:
        new_calendar = calendar.Calendar()
        month = list(new_calendar.itermonthdays(year, month))
        month_array = np.array(list(month))
        month_array = month_array.reshape(len(month_array) // 7, 7)
        return list(map(lambda x: list(x), filter(lambda x: any(x), month_array)))
    except BaseException as e:
        print(e)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Window()
    window.show()
    sys.exit(app.exec())
